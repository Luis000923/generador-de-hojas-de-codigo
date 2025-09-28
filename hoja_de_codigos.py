import pandas as pd
import numpy as np
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

class SurveyProcessor:
    def __init__(self):
        self.data = None
        self.questions = {}
        self.alternatives = {}
        self.sample_size = 0
        
    def load_excel(self, file_path):
        """Carga el archivo Excel con los datos de la encuesta"""
        try:
            # Leer el archivo Excel
            self.data = pd.read_excel(file_path, sheet_name=0)
            
            # Detectar automáticamente el tamaño de la muestra
            self.sample_size = len(self.data)
            
            # Identificar las columnas de preguntas (ignorar columnas administrativas)
            question_cols = [col for col in self.data.columns 
                           if not any(word in str(col).lower() 
                           for word in ['email', 'correo', 'mail', '@', 'grado', 'marca temporal', 'timestamp', 'fecha'])]
            
            # Procesar cada pregunta
            for i, col in enumerate(question_cols, 1):
                self.questions[f"Pregunta {i}"] = col
                # Obtener valores únicos como alternativas
                unique_values = self.data[col].dropna().unique()
                self.alternatives[f"Pregunta {i}"] = sorted(unique_values)
            
            return True
        except Exception as e:
            return f"Error al cargar el archivo: {str(e)}"
    
    def generate_coding_sheet(self):
        """Genera la hoja de código con el muestreo"""
        if self.data is None:
            return None
        
        coding_data = []
        
        for idx, row in self.data.iterrows():
            row_code = []
            row_code.append(f"No:___{idx + 1}___")
            
            for q_num, q_title in self.questions.items():
                # Obtener el número de pregunta
                q_idx = int(q_num.split()[1])
                value = row[q_title]
                
                # Codificar el valor
                if pd.notna(value):
                    alternatives = self.alternatives[q_num]
                    if value in alternatives:
                        code = alternatives.index(value) + 1
                    else:
                        code = 0  # Código para respuesta no válida
                else:
                    code = 0  # Código para sin respuesta
                
                row_code.append(f"{q_idx} {code}")
            
            coding_data.append(row_code)
        
        return coding_data
    
    def calculate_statistics(self):
        """Calcula estadísticas para cada pregunta"""
        statistics = {}
        
        for q_num, q_title in self.questions.items():
            # Contar frecuencias
            freq_table = self.data[q_title].value_counts()
            
            # Crear tabla de frecuencias con porcentajes
            stats_data = []
            total = len(self.data[q_title].dropna())
            
            for alternative in self.alternatives[q_num]:
                count = freq_table.get(alternative, 0)
                percentage = (count / total * 100) if total > 0 else 0
                stats_data.append({
                    'Alternativa': alternative,
                    'FA': count,
                    '%': f"{percentage:.1f}%"
                })
            
            statistics[q_num] = {
                'title': q_title,
                'data': stats_data,
                'total': total
            }
        
        return statistics
    
    def generate_excel_document(self, output_path, max_respondents=None):
        """Genera un archivo Excel con la hoja de código.
        max_respondents: si se indica (>0), limita la cantidad de encuestados exportados.
        """
        if self.data is None:
            return False
        
        # Crear un nuevo workbook
        wb = openpyxl.Workbook()
        
        # Eliminar la hoja por defecto
        wb.remove(wb.active)
        
        # ========== HOJA ÚNICA: SOLO HOJA DE CÓDIGO - MUESTREO ==========
        ws = wb.create_sheet("Hoja de Código - Muestreo")
        
        # Título de la hoja de código
        ws['A1'] = "Hoja de Código - Muestreo"
        ws['A1'].font = Font(bold=True, size=12)
        
        # Generar los datos de la hoja de código
        coding_data = self.generate_coding_sheet()

        # --- Bloque superior: por encuestado, una fila con celdas diagonales ---
        # Cada celda de pregunta contiene: número de pregunta (arriba-izquierda)
        # y código de respuesta (abajo-derecha), separados por una diagonal.
        current_row = 2
        if coding_data:
            total_questions = len(self.questions)
            num_q = min(5, max(0, total_questions))

            # Ajuste de anchos solo una vez
            ws.column_dimensions['A'].width = 6  # para el listado posterior
            for j in range(num_q):
                col_letter = openpyxl.utils.get_column_letter(2 + j)
                ws.column_dimensions[col_letter].width = 6

            # Determinar cuántos encuestados exportar
            total_rows = len(coding_data)
            limit = total_rows if not max_respondents or int(max_respondents) <= 0 else min(int(max_respondents), total_rows)

            for idx, row_vals in enumerate(coding_data[:limit], start=1):
                if num_q == 0:
                    break

                # Fila de encabezado (arriba): colocar "N° {idx}" encima del último recuadro de preguntas
                header_row = current_row
                header_col = 2 + (num_q - 1)
                header_cell = ws.cell(row=header_row, column=header_col)
                header_cell.value = f"N° {idx}"
                header_cell.alignment = Alignment(horizontal='center', vertical='center')
                header_cell.font = Font(bold=True)
                header_cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                            top=Side(style='thin'), bottom=Side(style='thin'))

                # Fila de preguntas (abajo): una celda por pregunta con diagonal y 2 números
                question_row = current_row + 1
                for j in range(num_q):
                    qnum = str(j + 1)
                    code_str = ""
                    if 1 + j < len(row_vals):
                        part = str(row_vals[1 + j])
                        parts = part.split()
                        code_str = parts[1] if len(parts) > 1 else part
                    self.set_diagonal_cell(ws, row=question_row, col=2 + j,
                                           top_left_text=qnum, bottom_right_text=code_str)

                # Ajuste de alto de fila para que se vea mejor la diagonal
                ws.row_dimensions[question_row].height = 22

                # Deja una fila en blanco y pasa al siguiente encuestado
                current_row = question_row + 2

        # Ya no escribimos el listado "No._i_ ..."; solo se generan los cuadritos por encuestado
        
        # Guardar el archivo Excel
        wb.save(output_path)
        return True
    
    def apply_diagonal_format(self, ws, cell_address, top_left_text, bottom_right_text):
        """
        Aplica un formato especial a una celda para simular la división diagonal
        """
        cell = ws[cell_address]
        
        # Crear el texto con formato que simule la división
        # Usamos caracteres especiales y espaciado para el efecto visual
        formatted_text = f"{top_left_text}{'':>8}\n{'':>6}{bottom_right_text}"
        
        cell.value = formatted_text
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(size=10)
        
        # Agregar bordes con diagonal
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin'),
            diagonal=Side(style='thin'),
            diagonalUp=True
        )
        
        return cell

    def set_diagonal_cell(self, ws, row, col, top_left_text, bottom_right_text):
        """Escribe en una celda un número arriba-izquierda y otro abajo-derecha
        simulando la separación con diagonal (borde diagonal + texto en dos líneas).
        """
        cell = ws.cell(row=row, column=col)
        # Composición en 2 líneas con espacios para acercar a las esquinas
        cell.value = f"{top_left_text}{'':>3}\n{'':>3}{bottom_right_text}"
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(size=11)
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
            diagonal=Side(style='thin'), diagonalUp=True
        )

class SurveyApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Generador de Hojas de Código")
        self.root.geometry("800x600")
        
        self.processor = SurveyProcessor()
        self.file_path = None
        
        self.setup_ui()
    
    def setup_ui(self):
        """Configura la interfaz de usuario"""
        # Frame principal
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Configurar el peso de las filas y columnas
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)

        # Título
        title_label = ttk.Label(main_frame, text="Generador de Hojas de Código", font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=3, pady=10)

        # Frame de carga de archivo
        file_frame = ttk.LabelFrame(main_frame, text="1. Cargar Archivo Excel", padding="10")
        file_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        self.file_label = ttk.Label(file_frame, text="No se ha seleccionado ningún archivo")
        self.file_label.grid(row=0, column=0, padx=5)
        ttk.Button(file_frame, text="Seleccionar Excel", command=self.load_file).grid(row=0, column=1, padx=5)

        # Frame de información de institución
        info_frame = ttk.LabelFrame(main_frame, text="2. Información de la Institución (Opcional)", padding="10")
        info_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        ttk.Label(info_frame, text="Institución:").grid(row=0, column=0, sticky=tk.W)
        self.institution_entry = ttk.Entry(info_frame, width=50)
        self.institution_entry.grid(row=0, column=1, padx=5)
        self.institution_entry.insert(0, "")
        ttk.Label(info_frame, text="Asignatura:").grid(row=1, column=0, sticky=tk.W)
        self.subject_entry = ttk.Entry(info_frame, width=50)
        self.subject_entry.grid(row=1, column=1, padx=5)
        ttk.Label(info_frame, text="Docente:").grid(row=2, column=0, sticky=tk.W)
        self.teacher_entry = ttk.Entry(info_frame, width=50)
        self.teacher_entry.grid(row=2, column=1, padx=5)

        # Frame de vista previa
        preview_frame = ttk.LabelFrame(main_frame, text="3. Vista Previa de Datos", padding="10")
        preview_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        main_frame.rowconfigure(3, weight=1)
        self.preview_text = scrolledtext.ScrolledText(preview_frame, height=10, width=70)
        self.preview_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        preview_frame.rowconfigure(0, weight=1)
        preview_frame.columnconfigure(0, weight=1)

        # Opciones de exportación
        options_frame = ttk.LabelFrame(main_frame, text="4. Opciones de Exportación", padding="10")
        options_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=6)
        ttk.Label(options_frame, text="N.º de encuestados a exportar (0 = todos):").grid(row=0, column=0, sticky=tk.W)
        self.max_resp_var = tk.IntVar(value=0)
        self.max_resp_spin = ttk.Spinbox(options_frame, from_=0, to=100000, width=8, textvariable=self.max_resp_var)
        self.max_resp_spin.grid(row=0, column=1, padx=8, sticky=tk.W)

        # Frame de botones de acción
        action_frame = ttk.Frame(main_frame)
        action_frame.grid(row=5, column=0, columnspan=3, pady=10)
        ttk.Button(action_frame, text="Generar Excel", command=self.generate_excel).grid(row=0, column=0, padx=5)
        ttk.Button(action_frame, text="Ver Estadísticas", command=self.show_statistics).grid(row=0, column=1, padx=5)
        ttk.Button(action_frame, text="Limpiar", command=self.clear_all).grid(row=0, column=2, padx=5)

        # Barra de estado
        self.status_bar = ttk.Label(main_frame, text="Listo", relief=tk.SUNKEN)
        self.status_bar.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E))
    
    def load_file(self):
        """Carga un archivo Excel"""
        self.file_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if self.file_path:
            result = self.processor.load_excel(self.file_path)
            
            if result == True:
                self.file_label.config(text=os.path.basename(self.file_path))
                self.status_bar.config(text=f"Archivo cargado: {self.processor.sample_size} registros")
                self.show_preview()
            else:
                messagebox.showerror("Error", result)
    
    def show_preview(self):
        """Muestra una vista previa de los datos cargados"""
        self.preview_text.delete(1.0, tk.END)
        
        preview = f"Muestra: {self.processor.sample_size} encuestados\n\n"
        preview += "PREGUNTAS DETECTADAS:\n"
        preview += "=" * 50 + "\n\n"
        
        for q_num, q_title in self.processor.questions.items():
            preview += f"{q_num}: {q_title}\n"
            alternatives = self.processor.alternatives[q_num]
            for i, alt in enumerate(alternatives, 1):
                preview += f"  {i}) {alt}\n"
            preview += "\n"
        
        self.preview_text.insert(1.0, preview)
    
    def show_statistics(self):
        """Muestra ventana con estadísticas"""
        if self.processor.data is None:
            messagebox.showwarning("Advertencia", "Primero debe cargar un archivo Excel")
            return
        
        # Crear ventana de estadísticas
        stats_window = tk.Toplevel(self.root)
        stats_window.title("Estadísticas de la Encuesta")
        stats_window.geometry("600x500")
        
        # Text widget para mostrar estadísticas
        stats_text = scrolledtext.ScrolledText(stats_window, height=25, width=70)
        stats_text.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
        
        # Generar y mostrar estadísticas
        statistics = self.processor.calculate_statistics()
        
        stats_display = f"ESTADÍSTICAS DE LA ENCUESTA\n"
        stats_display += f"Muestra Total: {self.processor.sample_size}\n"
        stats_display += "=" * 60 + "\n\n"
        
        for i, (q_num, stats) in enumerate(statistics.items(), 1):
            stats_display += f"CUADRO #{i}\n"
            stats_display += f"{stats['title']}\n"
            stats_display += f"N={self.processor.sample_size}\n"
            stats_display += "-" * 40 + "\n"
            stats_display += f"{'Alternativa':<30} {'FA':<10} {'%':<10}\n"
            stats_display += "-" * 40 + "\n"
            
            for row in stats['data']:
                stats_display += f"{str(row['Alternativa']):<30} {row['FA']:<10} {row['%']:<10}\n"
            
            stats_display += "-" * 40 + "\n"
            stats_display += f"{'Total':<30} {stats['total']:<10} {'100.0%':<10}\n"
            stats_display += "\n" + "=" * 60 + "\n\n"
        
        stats_text.insert(1.0, stats_display)
        stats_text.config(state=tk.DISABLED)
    
    def generate_excel(self):
        """Genera el archivo Excel con hoja de código y tablas"""
        if self.processor.data is None:
            messagebox.showwarning("Advertencia", "Primero debe cargar un archivo Excel")
            return
        
        # Solicitar ubicación para guardar
        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not output_path:
            return
        
        # Generar archivo Excel
        try:
            # Limitar el número de encuestados si se indicó en la UI
            max_resp = None
            try:
                max_resp = int(self.max_resp_var.get()) if self.max_resp_var.get() else None
            except Exception:
                max_resp = None
            self.processor.generate_excel_document(output_path, max_respondents=max_resp)
            self.status_bar.config(text=f"Archivo Excel generado: {os.path.basename(output_path)}")
            messagebox.showinfo("Éxito", f"Archivo Excel generado exitosamente:\n{output_path}")
            
            # Preguntar si desea abrir el archivo
            if messagebox.askyesno("Abrir archivo", "¿Desea abrir el archivo Excel generado?"):
                os.startfile(output_path)
        except Exception as e:
            messagebox.showerror("Error", f"Error al generar el archivo Excel: {str(e)}")
    
    def clear_all(self):
        """Limpia todos los datos"""
        self.processor = SurveyProcessor()
        self.file_path = None
        self.file_label.config(text="No se ha seleccionado ningún archivo")
        self.preview_text.delete(1.0, tk.END)
        self.status_bar.config(text="Listo")

def main():
    root = tk.Tk()
    app = SurveyApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()